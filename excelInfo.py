############################################################################################
# Creates the excel file that will be used to hold information. This will allow the 
# program to compare new data against old, and also allow admins to make changes to
# the data on an "as-needed" basis
############################################################################################

from openpyxl import *
import csv

# Dictionary/Constants
BOOK = 'IB_Tracker.xlsx'

def writeInfo(clanList):
    ############################################################################################
    # Saves the data in the spreadsheet 
    ############################################################################################

    # Open the book
    wb = load_workbook(BOOK)
    ws = wb.get_sheet_by_name("IB")

    # Write the info to the spreadsheet
    r = 1
    rr = 1
    c = 1
    cc = 3
    for i in clanList:
        ws.cell(row=r, column=c).value=i.displayName
        ws.cell(row=r,column=c+1).value=i.memberID
        for char in i.memberChars:
            ws.cell(row=rr, column=cc).value=char.values
            rr += 1
            cc += 1 
        r += 1
        
    wb.save    

def writeCSV(clanList):
    ############################################################################################
    # Saves the data to a CSV file, because spreadsheets are hard.... 
    ############################################################################################

    with open('ib.csv', 'w', newline='') as csvfile:
        headers = ['Username', 'MemberID', 'Char Info']
        clanWriter = csv.writer(csvfile, delimiter=' ',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)

        #clanWriter.writerow(headers)
        #clanWriter.writeheader()
        for i in clanList:
            clanWriter.writerow([i])